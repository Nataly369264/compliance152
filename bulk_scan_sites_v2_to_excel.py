"""Bulk scan of 20 sites → fill columns F and G of Лист 3."""
import asyncio
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from src.scanner.crawler import SiteScanner
from src.scanner.tracker_registry import find_trackers_in_scripts

EXCEL_PATH = "/Users/dianayagubov/Downloads/Золотой_набор_заполнил.xlsx"

SITES = [
    "skyeng.ru",
    "geekbrains.ru",
    "foxford.ru",
    "skillbox.ru",
    "netology.ru",
    "stepik.org",
    "coursera.org",
    "uchi.ru",
    "rosuchebnik.ru",
    "profil-edu.ru",
    "egecrm.ru",
    "repetitors.info",
    "100ballov.ru",
    "videouroki.net",
    "infourok.ru",
    "yandex.ru",
    "sferum.ru",
    "school-primer.ru",
    "учи.рф",
    "учитель.com",
]

# 37 checks per site, start row = 3 + site_index * 37
# col F = 6 (result), col G = 7 (notes)

def map_to_37(result) -> list[tuple[str, str]]:
    """Return list of 37 (value, note) tuples for all checks."""
    pp = result.privacy_policy
    cb = result.cookie_banner
    pd_forms = [f for f in result.forms if f.collects_personal_data]
    all_scripts = result.external_scripts
    service_names = {s.service_name for s in all_scripts if s.service_name}
    domains = [s.domain for s in all_scripts if s.domain]
    trackers = find_trackers_in_scripts(domains)
    foreign_trackers = [t for t in trackers if t["is_foreign"]]

    def yn(val): return "Да" if val else "Нет"
    def det(val): return "Обнаружен" if val else "Не обнаружен"

    # ── ФОРМЫ ──────────────────────────────────────────────────
    if pd_forms:
        f1 = (yn(all(f.has_consent_checkbox for f in pd_forms)),
              f"{len(pd_forms)} форм с ПДн; без согласия: {sum(1 for f in pd_forms if not f.has_consent_checkbox)}")
        f2 = (yn(not any(f.consent_checkbox_prechecked for f in pd_forms)),
              "Предотмеченных чекбоксов: " + str(sum(1 for f in pd_forms if f.consent_checkbox_prechecked)))
        f3 = (yn(all(f.has_privacy_link for f in pd_forms)),
              f"Форм без ссылки на политику: {sum(1 for f in pd_forms if not f.has_privacy_link)}")
        f4 = (yn(any(f.has_marketing_checkbox for f in pd_forms)),
              "Маркетинговый чекбокс " + ("найден" if any(f.has_marketing_checkbox for f in pd_forms) else "не найден"))
        max_f = max(len(f.personal_data_fields) for f in pd_forms)
        f5 = (yn(max_f <= 5),
              f"Макс. полей ПДн в одной форме: {max_f}")
    else:
        f1 = ("?", "Форм, собирающих ПДн, не обнаружено — статический сканер мог не увидеть JS-формы")
        f2 = ("?", "Форм не обнаружено")
        f3 = ("?", "Форм не обнаружено")
        f4 = ("?", "Форм не обнаружено")
        f5 = ("?", "Форм не обнаружено")

    # ── COOKIE ─────────────────────────────────────────────────
    if cb.found:
        c1 = ("Да", "Cookie-баннер обнаружен статическим парсером")
        c2 = (yn(cb.has_decline_button),
              "Кнопка отклонения " + ("есть" if cb.has_decline_button else "не найдена"))
        c3 = (yn(cb.has_category_choice),
              "Выбор категорий " + ("есть" if cb.has_category_choice else "не найден"))
        c4 = (yn(not cb.analytics_before_consent),
              "Аналитика до согласия: " + ("да — нарушение" if cb.analytics_before_consent else "не обнаружено"))
    else:
        c1 = ("?", "Баннер не найден статическим сканером. Если сайт на JS-фреймворке (React/Next.js) — проверить вручную в режиме инкогнито")
        c2 = ("?", "Баннер не обнаружен — проверить вручную")
        c3 = ("?", "Баннер не обнаружен — проверить вручную")
        c4 = ("?", "Проверить через DevTools → Network до нажатия на баннер")

    # ── ПОЛИТИКА ───────────────────────────────────────────────
    if not pp.found:
        note_pp = "Политика не найдена статическим сканером. На JS-сайтах ссылка рендерится через JS — проверить вручную"
        p = [("?", note_pp)] * 17
    elif not pp.text:
        p = [("?", "Политика найдена как PDF или текст недоступен — содержимое не анализируется, проверить вручную")]
        p += [("?", "Текст политики недоступен (PDF или пусто) — проверить вручную")] * 16
        p[0] = ("Да", f"Политика найдена: {pp.url} (без текста — вероятно PDF)")
        p[1] = (yn(pp.in_footer), "Ссылка в футере: " + ("найдена" if pp.in_footer else "не найдена"))
    else:
        p = []
        p.append(("Да", f"Политика найдена: {pp.url}"))
        p.append((yn(pp.in_footer), "Ссылка в футере: " + ("найдена" if pp.in_footer else "не найдена на всех страницах")))
        p.append((yn(pp.has_operator_name), "Наименование оператора " + ("найдено" if pp.has_operator_name else "не найдено в тексте")))
        p.append((yn(pp.has_inn_ogrn), "ИНН/ОГРН " + ("найдены" if pp.has_inn_ogrn else "не найдены в тексте")))
        p.append((yn(pp.has_responsible_person), "Контакт ответственного " + ("найден" if pp.has_responsible_person else "не найден")))
        p.append((yn(pp.has_data_categories), "Категории ПДн " + ("перечислены" if pp.has_data_categories else "не перечислены")))
        p.append((yn(pp.has_purposes), "Цели обработки " + ("указаны" if pp.has_purposes else "не указаны")))
        p.append((yn(pp.has_legal_basis), "Правовые основания " + ("указаны" if pp.has_legal_basis else "не указаны")))
        p.append((yn(pp.has_retention_periods), "Сроки хранения " + ("указаны" if pp.has_retention_periods else "не указаны")))
        p.append((yn(pp.has_subject_rights), "Права субъектов " + ("упомянуты" if pp.has_subject_rights else "не упомянуты")))
        p.append((yn(pp.has_rights_procedure), "Порядок реализации прав " + ("описан" if pp.has_rights_procedure else "не описан")))
        p.append((yn(pp.has_cross_border_info), "Трансграничная передача " + ("упомянута" if pp.has_cross_border_info else "не упомянута")))
        p.append((yn(pp.has_security_measures), "Меры безопасности " + ("описаны" if pp.has_security_measures else "не описаны")))
        p.append((yn(pp.has_cookie_info), "Информация о cookies " + ("есть" if pp.has_cookie_info else "отсутствует")))
        p.append((yn(pp.has_localization_statement), "Локализация в РФ " + ("упомянута" if pp.has_localization_statement else "не упомянута")))
        p.append((yn(pp.has_date), "Дата публикации " + ("есть" if pp.has_date else "не найдена")))
        p.append((yn(pp.is_russian), "Язык документа: " + ("русский" if pp.is_russian else "не определён как русский")))

    # ── ТЕХНИЧЕСКИЕ ────────────────────────────────────────────
    t_ssl = (yn(result.ssl_info.has_ssl), "HTTPS " + ("активен" if result.ssl_info.has_ssl else "не настроен — нарушение"))
    t_fonts = (det("Google Fonts" in service_names),
               ("Запрос к fonts.googleapis.com обнаружен — нарушение с 01.07.2025" if "Google Fonts" in service_names else "Запросы к Google Fonts не обнаружены"))
    t_ga = (det("Google Analytics" in service_names),
            ("Google Analytics обнаружен — нарушение с 01.07.2025" if "Google Analytics" in service_names else "Google Analytics не обнаружен"))
    fb_vk = "Facebook Pixel" in service_names or "VK Pixel" in service_names or "myTarget" in service_names
    t_pixel = (det(fb_vk),
               ("Facebook/VK Pixel обнаружен — нарушение" if fb_vk else "Facebook Pixel и VK Pixel не обнаружены"))
    t_captcha = (det("Google reCAPTCHA" in service_names),
                 ("Google reCAPTCHA обнаружена — нарушение с 01.07.2025" if "Google reCAPTCHA" in service_names else "Google reCAPTCHA не обнаружена"))
    t_gtm = (det("Google Tag Manager" in service_names),
             ("Google Tag Manager обнаружен — нарушение с 01.07.2025" if "Google Tag Manager" in service_names else "Google Tag Manager не обнаружен"))

    # ── ТРЕКЕРЫ ────────────────────────────────────────────────
    if not pp.found or not pp.text:
        tr1 = ("?", f"Трекеры на сайте: {', '.join(t['name'] for t in trackers) or 'нет'}. Текст политики недоступен — проверить упоминание вручную")
        tr2 = ("?", f"Иностранные трекеры: {', '.join(t['name'] for t in foreign_trackers) or 'нет'}. Проверить раздел о трансграничной передаче в политике вручную")
    else:
        policy_text = pp.text.lower()
        undisclosed = [t for t in trackers if not any(kw in policy_text for kw in t["keywords"])]
        tr1 = (yn(not undisclosed),
               f"Трекеры на сайте: {', '.join(t['name'] for t in trackers) or 'нет'}. " +
               (f"Не упомянуты в политике: {', '.join(t['name'] for t in undisclosed)}" if undisclosed else "Все упомянуты"))
        has_cross = any(kw in policy_text for kw in ["трансграничн", "передача за рубеж", "иностранн", "третьи страны", "зарубежн"])
        foreign_undisclosed = [t for t in foreign_trackers if not has_cross]
        tr2 = (yn(not foreign_undisclosed),
               f"Иностранные трекеры: {', '.join(t['name'] for t in foreign_trackers) or 'нет'}. " +
               ("Трансграничная передача раскрыта в политике" if not foreign_undisclosed else "Раскрытие трансграничной передачи отсутствует"))

    # ── РЕГУЛЯТОРНЫЕ ───────────────────────────────────────────
    reg = [
        ("?", "Требует ручной проверки: pd.rkn.gov.ru → Реестр операторов → поиск по ИНН"),
        ("?", "Требует ручной проверки: pd.rkn.gov.ru → Реестр операторов → уведомление"),
        ("?", "Проверить через 2ip.ru — страна сервера должна быть RU (Россия)"),
    ]

    result_37 = [f1, f2, f3, f4, f5,
                 c1, c2, c3, c4,
                 *p,
                 t_ssl, t_fonts, t_ga, t_pixel, t_captcha, t_gtm,
                 tr1, tr2,
                 *reg]
    return result_37


async def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws3 = wb.worksheets[2]

    scanner = SiteScanner(max_pages=5, timeout=20, crawl_delay=0.5)

    # skyeng уже заполнен вручную — начинаем с индекса 1 (geekbrains)
    for site_idx, site in enumerate(SITES):
        if site_idx == 0:
            print(f"[01/20] skyeng.ru — уже заполнен, пропускаем")
            continue

        print(f"\n[{site_idx+1:02d}/20] Scanning {site} ...", flush=True)
        try:
            result = await scanner.scan(site)
            checks = map_to_37(result)
            pp = result.privacy_policy
            print(f"       pages={result.pages_scanned}  ssl={result.ssl_info.has_ssl}  "
                  f"policy={pp.found}  cookie={result.cookie_banner.found}  "
                  f"forms_pd={sum(1 for f in result.forms if f.collects_personal_data)}")
        except Exception as exc:
            print(f"       ERROR: {exc}")
            checks = [("?", f"Ошибка сканирования: {exc}")] * 37

        start_row = 3 + site_idx * 37
        for i, (val, note) in enumerate(checks):
            ws3.cell(row=start_row + i, column=6, value=val)
            ws3.cell(row=start_row + i, column=7, value=note)

        wb.save(EXCEL_PATH)
        print(f"       Сохранено → строки {start_row}–{start_row+36}")

    print("\n✓ Готово. Файл сохранён:", EXCEL_PATH)


if __name__ == "__main__":
    asyncio.run(main())
