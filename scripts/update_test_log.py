#!/usr/bin/env python3
"""
update_test_log.py — добавляет прогон из golden run JSON в scanner_test_log.xlsx.

Использование:
    python scripts/update_test_log.py <golden_run.json> [scanner_test_log.xlsx]

Алгоритм:
  1. Читает JSON прогона и ODS-эталон (golden_set_v1.ods)
  2. Находит следующий run_id
  3. Добавляет строку на лист Прогоны
  4. Добавляет строки на лист Результаты (все 37 эталонных пунктов + extra-checks)
  5. Создаёт задачи на листе Задачи для каждого РАСХОЖДЕНИЯ (без дублей)
  6. Сохраняет файл
"""

import sys, io, json, pathlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

# =============================================
# КОНСТАНТЫ: маппинг эталон 1-37 → CHECK_ID
# =============================================

# (item_id: str, item_name: str, check_id: str | None, inverted: bool)
# inverted=True для TECH_002..006: в эталоне Обнаружен=нарушение,
# в сканере FAIL=найдено (т.е. FAIL↔Обнаружен, PASS↔Не обнаружен)
CHECKLIST_MAP = [
    ('1',  'Согласие на обработку ПД в форме',          'FORM_001',   False),
    ('2',  'Чекбокс согласия не предотмечен',           'FORM_002',   False),
    ('3',  'Ссылка на политику рядом с формой',         'FORM_003',   False),
    ('4',  'Отдельный чекбокс для маркетинга',          'FORM_006',   False),
    ('5',  'Минимальность собираемых данных',           'FORM_007',   False),
    ('6',  'Cookie-баннер присутствует',                'COOKIE_001', False),
    ('7',  'Кнопка «Отклонить» в баннере',              'COOKIE_002', False),
    ('8',  'Выбор категорий cookie',                    'COOKIE_003', False),
    ('9',  'Аналитика не грузится до согласия',         'COOKIE_005', False),
    ('10', 'Политика опубликована',                     'POLICY_001', False),
    ('11', 'Ссылка на политику в футере',               'POLICY_002', False),
    ('12', 'Полное наименование оператора',             'POLICY_003', False),
    ('13', 'ИНН / ОГРН оператора',                     'POLICY_004', False),
    ('14', 'Контакт ответственного за ПДн',             'POLICY_005', False),
    ('15', 'Категории персональных данных',             'POLICY_006', False),
    ('16', 'Цели обработки',                            'POLICY_007', False),
    ('17', 'Правовые основания обработки',              'POLICY_008', False),
    ('18', 'Сроки хранения данных',                     'POLICY_009', False),
    ('19', 'Права субъектов ПДн',                      'POLICY_010', False),
    ('20', 'Порядок реализации прав (10 р. дней)',      'POLICY_011', False),
    ('21', 'Информация о трансграничной передаче',      'POLICY_012', False),
    ('22', 'Описание мер безопасности',                 'POLICY_013', False),
    ('23', 'Информация о cookies в политике',           'POLICY_014', False),
    ('24', 'Локализация данных в РФ',                   'POLICY_015', False),
    ('25', 'Дата публикации и обновления',              'POLICY_016', False),
    ('26', 'Документ на русском языке',                 'POLICY_017', False),
    ('27', 'SSL / HTTPS',                               'TECH_001',   False),
    ('28', 'Google Fonts',                              'TECH_002',   True),
    ('29', 'Google Analytics',                          'TECH_003',   True),
    ('30', 'Facebook Pixel / VK Pixel',                 'TECH_004',   True),
    ('31', 'Google reCAPTCHA',                          'TECH_005',   True),
    ('32', 'Google Tag Manager',                        'TECH_006',   True),
    ('33', 'Трекеры упомянуты в политике ПДн',         'TRACKER_001',False),
    ('34', 'Трансграничная передача раскрыта',          None,         False),
    ('35', 'Оператор в реестре РКН',                    'REG_001',    False),
    ('36', 'Уведомление подано в РКН',                  'REG_002',    False),
    ('37', 'Хостинг на территории РФ',                  'REG_003',    False),
]


def normalize_ods_value(raw: str, item_id: str) -> str:
    """Нормализует значение из golden_set_v1.ods в формат эталона."""
    r = raw.strip()
    # Технические пункты 28-32: ✅=Обнаружен, ❌=Не обнаружен
    if item_id in ('28', '29', '30', '31', '32'):
        if '✅' in r:    return 'Обнаружен'
        if '❌' in r:    return 'Не обнаружен'
        if r:            return r
        return 'нет данных'
    # Остальные пункты
    if '✅' in r:        return 'Да'
    if '❌' in r:        return 'Нет'
    if '⚠' in r:        return 'Частично(неявное)'
    if r in ('Предотмечен',): return 'Нет'
    if r in ('Грузится',):    return 'Нет'
    if r.startswith('Не указан'):  return 'Нет'
    if r.startswith('Указано, что не осуществляется'): return 'Частично(неявное)'
    if r == '':          return 'нет данных'
    return r


def scanner_to_result(status: str) -> str:
    """Переводит статус сканера в PASS / FAIL / не проверялось."""
    s = status.lower()
    if s in ('pass',):                       return 'PASS'
    if s in ('fail', 'warning'):             return 'FAIL'
    return 'не проверялось'


def compare(scanner_result: str, etalon: str) -> str:
    """Вычисляет значение колонки 'сравнение'.

    Учитывает инвертированную семантику TECH-пунктов:
      FAIL + Обнаружен → ✓ (сканер нашёл сервис = совпадает с эталоном)
      PASS + Не обнаружен → ✓ (сервис отсутствует = совпадает с эталоном)
    """
    if scanner_result == 'не проверялось' or etalon in ('нет данных', ''):
        return 'НУЖНА ПРОВЕРКА'
    s, e = scanner_result, etalon
    if s == 'PASS' and e in ('Да', 'Частично(неявное)', 'Не обнаружен'):
        return '✓'
    if s == 'FAIL' and e in ('Нет', 'Обнаружен'):
        return '✓'
    return 'РАСХОЖДЕНИЕ'


def classify_gap(scanner_result: str, etalon: str):
    """Возвращает (тип_расхождения, серьёзность) для РАСХОЖДЕНИЯ."""
    s, e = scanner_result, etalon
    # Сканер пропустил нарушение: PASS когда должен FAIL
    if s == 'PASS' and e in ('Нет', 'Обнаружен'):
        return 'false negative', 'критично'
    # Сканер ложно сработал: FAIL когда всё ок
    if s == 'FAIL' and e in ('Да', 'Частично(неявное)', 'Не обнаружен'):
        return 'false positive', 'важно'
    return 'расхождение', 'важно'


def gap_hypothesis(gap_type: str, item_name: str) -> str:
    if gap_type == 'false negative':
        return (f'Сканер не распознаёт признак нарушения для пункта «{item_name}». '
                'Возможно: неверный селектор, логика проверки не покрывает этот кейс.')
    if gap_type == 'false positive':
        return (f'Сканер ложно срабатывает на пункте «{item_name}». '
                'Возможно: слишком широкое условие, неверная интерпретация элемента.')
    return 'Требует ручного анализа.'


# =============================================
# ЧИТАЕМ АРГУМЕНТЫ
# =============================================

args = sys.argv[1:]
if not args:
    print('Использование: python scripts/update_test_log.py <golden_run.json> [scanner_test_log.xlsx]')
    sys.exit(1)

golden_json_path = pathlib.Path(args[0])
log_path         = pathlib.Path(args[1] if len(args) > 1 else 'scanner_test_log.xlsx')
ods_path         = pathlib.Path('tests/fixtures/golden_set_v1.ods')

# =============================================
# ЧИТАЕМ GOLDEN RUN JSON
# =============================================

data   = json.loads(golden_json_path.read_text(encoding='utf-8'))
ri     = data.get('run_info', {})
report = data['compliance_report']

site_url    = ri.get('url', '').replace('https://', '').replace('http://', '').rstrip('/')
run_at_str  = ri.get('run_at', '')
score       = report.get('overall_score', 0)
scanner_ver = ri.get('scanner', 'SiteScanner')

# Дата_время прогона
try:
    run_dt = datetime.fromisoformat(run_at_str.replace('Z', '+00:00'))
    dt_str = run_dt.strftime('%Y-%m-%d %H:%M')
except Exception:
    dt_str = run_at_str[:16]

# Строим словарь check_id -> scanner_result (PASS/FAIL/не проверялось)
checklist_raw   = {item['id']: item['status'] for item in report.get('checklist', [])}
scanner_results = {cid: scanner_to_result(st) for cid, st in checklist_raw.items()}

# =============================================
# ЧИТАЕМ ОДС-ЭТАЛОН (el-ed.ru / аналогичный)
# =============================================

ods_etalon = {}   # item_id -> normalized_value
ods_date   = '2025-04-27'
ods_author = 'Наталья'

try:
    import odf.opendocument as oo, odf.table as ot, odf.text as ox
    doc  = oo.load(str(ods_path))
    sh   = doc.spreadsheet.getElementsByType(ot.Table)[0]
    rows = sh.getElementsByType(ot.TableRow)

    def cell_val(c):
        ps = c.getElementsByType(ox.P)
        if ps and ps[0].firstChild:
            return ps[0].firstChild.data
        return ''

    for row in rows:
        cells = row.getElementsByType(ot.TableCell)
        vals  = [cell_val(c) for c in cells]
        if len(vals) >= 5 and vals[1].strip().isdigit():
            iid = vals[1].strip()
            raw = vals[4].strip() if len(vals) > 4 else ''
            ods_etalon[iid] = normalize_ods_value(raw, iid)
except Exception as e:
    print(f'⚠  Не удалось прочитать ODS: {e}. Эталон будет "нет данных".')

# =============================================
# ОТКРЫВАЕМ ФАЙЛ ЛОГА
# =============================================

if not log_path.exists():
    print(f'Файл {log_path} не найден. Сначала запусти scripts/create_test_log.py')
    sys.exit(1)

wb = openpyxl.load_workbook(str(log_path))
ws_runs   = wb['Прогоны']
ws_res    = wb['Результаты']
ws_tasks  = wb['Задачи']
ws_etalon = wb['Эталон']

# =============================================
# Определяем следующий run_id и task_id
# =============================================

def next_id(ws, prefix, col=1):
    max_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[col - 1]
        if v and str(v).startswith(prefix):
            try:
                n = int(str(v)[len(prefix):])
                if n > max_num:
                    max_num = n
            except ValueError:
                pass
    return f'{prefix}{max_num + 1:03d}'

run_id  = next_id(ws_runs, 'RUN_')
task_id_base = int(next_id(ws_tasks, 'BUG_')[4:])

# =============================================
# ШАГ 1: Читаем эталон из листа Эталон (xlsx) — приоритет над ODS
# =============================================

# Строим словарь эталонных значений из уже имеющегося листа Эталон
# Формат ключа в листе: "site|item_id" (колонка A)
existing_keys = set()
sheet_etalon  = {}   # (site, item_id) -> value из листа Эталон
for row in ws_etalon.iter_rows(min_row=2, values_only=True):
    if row[0]:
        existing_keys.add(str(row[0]))
    if row[1] and row[2] and row[4] is not None:
        sheet_etalon[(str(row[1]), str(row[2]))] = str(row[4])

def get_etalon(site: str, iid: str) -> str:
    """Возвращает эталонное значение: сначала из листа Эталон, затем из ODS."""
    return sheet_etalon.get((site, iid), ods_etalon.get(iid, 'нет данных'))

# Добавляем строки в Эталон только для новых сайтов/пунктов
etalon_added = 0
for iid, iname, cid, inverted in CHECKLIST_MAP:
    key = f'{site_url}|{iid}'
    if key not in existing_keys:
        val = ods_etalon.get(iid, 'нет данных')
        ws_etalon.append([key, site_url, iid, iname, val, ods_date, ods_author])
        sheet_etalon[(site_url, iid)] = val
        etalon_added += 1

# =============================================
# ШАГ 2: Добавляем строку в Прогоны
# =============================================

def pct_formula(r: int) -> str:
    return (
        f'=IFERROR('
        f'COUNTIFS(Результаты!$A:$A,A{r},Результаты!$G:$G,"✓")'
        f'/(COUNTIFS(Результаты!$A:$A,A{r},Результаты!$G:$G,"✓")'
        f'+COUNTIFS(Результаты!$A:$A,A{r},Результаты!$G:$G,"РАСХОЖДЕНИЕ"))'
        f'*100,0)'
    )

next_run_row = ws_runs.max_row + 1
ws_runs.append([
    run_id,
    dt_str,
    site_url,
    f'{score}%',
    pct_formula(next_run_row),
    '—',          # время_сек — нет в JSON
    scanner_ver,
    f'golden run {golden_json_path.name}'
])

# =============================================
# ШАГ 3: Добавляем строки в Результаты
# =============================================

def f_etalon(r: int) -> str:
    return f'=IFERROR(VLOOKUP(B{r}&"|"&C{r},Эталон!$A:$E,5,0),"нет данных")'

def f_srav(r: int) -> str:
    e, f = f'E{r}', f'F{r}'
    # PASS+Да/Частично/Не_обнаружен → ✓
    # FAIL+Нет/Обнаружен → ✓
    # Остальное → РАСХОЖДЕНИЕ
    return (
        f'=IF(OR({e}="не проверялось",{f}="нет данных",{f}=""),"НУЖНА ПРОВЕРКА",'
        f'IF(OR('
        f'AND({e}="PASS",OR({f}="Да",{f}="Частично(неявное)",{f}="Не обнаружен")),'
        f'AND({e}="FAIL",OR({f}="Нет",{f}="Обнаружен"))),'
        f'"✓","РАСХОЖДЕНИЕ"))'
    )

task_rows   = []   # задачи для создания
расхождения = []   # для сводки

# 37 эталонных пунктов
for iid, iname, cid, inverted in CHECKLIST_MAP:
    scanner_res = scanner_results.get(cid, 'не проверялось') if cid else 'не проверялось'
    etalon_val  = get_etalon(site_url, iid)   # приоритет: лист Эталон → ODS

    comp = compare(scanner_res, etalon_val)
    disc_type, severity, comment = '', '', ''

    if comp == 'РАСХОЖДЕНИЕ':
        disc_type, severity = classify_gap(scanner_res, etalon_val)
        comment = (f'Сканер: {scanner_res}, эталон: {etalon_val}')
        расхождения.append((iid, iname, disc_type, severity, scanner_res, etalon_val))

    rn = ws_res.max_row + 1
    ws_res.append([
        run_id, site_url, iid, iname,
        scanner_res,
        f_etalon(rn),
        f_srav(rn),
        disc_type, severity, comment
    ])

# Дополнительные пункты сканера (не в эталоне 1-37)
EXTRA_CHECKS = ['CONSENT_001', 'CONSENT_002', 'CONSENT_003', 'CONSENT_004',
                'CONSENT_005', 'TRACKER_002']
EXTRA_NAMES  = {
    'CONSENT_001': 'Наличие согласия на обработку ПДн',
    'CONSENT_002': 'Согласие отделено от оферты',
    'CONSENT_003': 'Обязательные реквизиты в тексте согласия',
    'CONSENT_004': 'Возможность отзыва согласия',
    'CONSENT_005': 'Раздельные согласия для разных целей',
    'TRACKER_002': 'Трекер без согласия пользователя',
}
for cid in EXTRA_CHECKS:
    if cid in scanner_results:
        scanner_res = scanner_results[cid]
        iname = EXTRA_NAMES.get(cid, cid)
        disc_type = 'новая находка' if scanner_res == 'FAIL' else ''
        severity  = 'минор' if scanner_res == 'FAIL' else ''
        rn = ws_res.max_row + 1
        ws_res.append([
            run_id, site_url, cid, iname,
            scanner_res,
            f_etalon(rn),  # вернёт "нет данных" — OK
            f_srav(rn),
            disc_type, severity,
            'Пункт отсутствует в эталоне 1–37, требует ручной верификации' if disc_type else ''
        ])

# =============================================
# ШАГ 4: Создаём задачи (без дублей)
# =============================================

# Собираем существующие открытые задачи site+item
open_tasks = set()
for row in ws_tasks.iter_rows(min_row=2, values_only=True):
    if row[9] == 'открыта' and row[2] and row[3]:  # статус, сайт, пункт_id
        open_tasks.add((str(row[2]), str(row[3])))

today_str = datetime.now().strftime('%Y-%m-%d')
new_tasks  = 0
skipped    = 0

for iid, iname, disc_type, severity, scanner_res, etalon_val in расхождения:
    key = (site_url, iid)
    if key in open_tasks:
        skipped += 1
        continue
    tid = f'BUG_{task_id_base:03d}'
    task_id_base += 1
    desc = (f'Сканер выдал {scanner_res}, эталон ожидает {etalon_val}. '
            f'Пункт: {iname}. Сайт: {site_url}')
    hyp = gap_hypothesis(disc_type, iname)
    ws_tasks.append([
        tid, run_id, site_url, iid, iname,
        disc_type, severity, desc, hyp,
        'открыта', today_str, '', ''
    ])
    open_tasks.add(key)
    new_tasks += 1

# =============================================
# ШАГ 5: Сохраняем
# =============================================

wb.save(str(log_path))

# =============================================
# ИТОГОВАЯ СВОДКА
# =============================================

all_comps   = [compare(
                   scanner_results.get(cid, 'не проверялось') if cid else 'не проверялось',
                   get_etalon(site_url, iid)
               ) for iid, iname, cid, inverted in CHECKLIST_MAP]

cnt_ok      = all_comps.count('✓')
cnt_gap     = all_comps.count('РАСХОЖДЕНИЕ')
cnt_check   = all_comps.count('НУЖНА ПРОВЕРКА')
total_open  = sum(1 for row in ws_tasks.iter_rows(min_row=2, values_only=True)
                  if row[9] == 'открыта')
crit_open   = sum(1 for row in ws_tasks.iter_rows(min_row=2, values_only=True)
                  if row[9] == 'открыта' and row[6] == 'критично')
imp_open    = sum(1 for row in ws_tasks.iter_rows(min_row=2, values_only=True)
                  if row[9] == 'открыта' and row[6] == 'важно')

print(f'✓ Прогон {run_id} добавлен в {log_path}')
print(f'  Сайт: {site_url}   Score сканера: {score}%   Дата: {dt_str}')
print()
print(f'  Результаты по эталону (из 37 пунктов):')
print(f'    ✓  совпадений:       {cnt_ok}')
print(f'    РАСХОЖДЕНИЕ:         {cnt_gap}')
print(f'    НУЖНА ПРОВЕРКА:      {cnt_check}')
if cnt_ok + cnt_gap > 0:
    pct = round(cnt_ok / (cnt_ok + cnt_gap) * 100, 1)
    print(f'    Точность (✓/сравн.): {pct}%')
print()
print(f'  Задачи:')
print(f'    Создано новых:       {new_tasks}')
print(f'    Пропущено (есть откр.): {skipped}')
print(f'    Всего открытых:      {total_open} '
      f'(🔴 критичных: {crit_open}, 🟡 важных: {imp_open})')
print()

if расхождения:
    print('  Расхождения:')
    for iid, iname, disc_type, severity, sr, ev in расхождения:
        icon = '🔴' if severity == 'критично' else '🟡'
        print(f'    {icon} [{iid}] {iname}')
        print(f'       сканер={sr}, эталон={ev} → {disc_type}')
