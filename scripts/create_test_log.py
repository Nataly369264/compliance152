#!/usr/bin/env python3
"""Create scanner_test_log.xlsx — QA-лог для сканера 152-ФЗ"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# =============================================
# ШАГ 1: Читаем эталонный файл
# =============================================

wb_src = openpyxl.load_workbook('сканер_152фз_v7_27.04.xlsx', data_only=True)
ws2 = wb_src['Лист 2 — Разметка']

sites_row = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
annotators_row = list(ws2.iter_rows(min_row=3, max_row=3, values_only=True))[0]

# Уникальные сайты в порядке появления
seen_sites = set()
site_cols = []  # [(col_index, site, annotator)]
for i, (site, ann) in enumerate(zip(sites_row, annotators_row)):
    if site and site != 'Проверка' and site not in seen_sites:
        seen_sites.add(site)
        site_cols.append((i, str(site), str(ann) if ann else 'неизвестно'))

# Пункты чеклиста и эталонные данные
checklist = []   # [(item_id, item_name, group)]
ref_data  = {}   # (site, item_id) -> raw_value
current_group = ''

for row in ws2.iter_rows(min_row=4, max_row=200, values_only=True):
    val = row[0]
    if val is None:
        continue
    s = str(val)
    if s.startswith('▌'):
        current_group = s[2:].strip()
        continue
    if s.startswith('Да/'):
        continue
    parts = s.split('. ', 1)
    if len(parts) == 2 and parts[0].strip().isdigit():
        item_id   = parts[0].strip()
        item_name = parts[1].strip()
        checklist.append((item_id, item_name, current_group))
        for col_i, site, ann in site_cols:
            if col_i < len(row) and row[col_i] is not None:
                ref_data[(site, item_id)] = str(row[col_i])

# =============================================
# ШАГ 2: Стили и утилиты
# =============================================

BOLD       = Font(bold=True)
GRAY_FILL  = PatternFill(fill_type='solid', fgColor='BDC3C7')
LGRAY_FILL = PatternFill(fill_type='solid', fgColor='E0E0E0')


def style_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = BOLD
        cell.fill = GRAY_FILL
        cell.alignment = Alignment(wrap_text=False)


def freeze_and_autofit(ws, narrow_a=False):
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if narrow_a and col_letter == 'A':
            ws.column_dimensions['A'].width = 3
            continue
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=8
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def cond_fmt(ws, cell_range, formula, bg_hex):
    fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[formula], fill=fill))


# =============================================
# ШАГ 3: Создаём книгу
# =============================================

wb = Workbook()

# ====================================================
# ЛИСТ 1: Прогоны
# ====================================================

ws_runs = wb.active
ws_runs.title = 'Прогоны'

style_header(ws_runs, [
    'run_id', 'дата_время', 'сайт', 'score_сканера',
    'совпадений_с_эталоном_%', 'время_сек', 'версия_сканера', 'примечания'
])


def pct_formula(r: int) -> str:
    """Процент совпадений с эталоном по run_id в строке r листа Прогоны."""
    return (
        f'=IFERROR('
        f'COUNTIFS(Результаты!$A:$A,A{r},Результаты!$G:$G,"✓")'
        f'/(COUNTIFS(Результаты!$A:$A,A{r},Результаты!$G:$G,"✓")'
        f'+COUNTIFS(Результаты!$A:$A,A{r},Результаты!$G:$G,"РАСХОЖДЕНИЕ"))'
        f'*100,0)'
    )


ws_runs.append(['RUN_001', '2026-05-01 10:00', 'skyeng.ru', '68%',
                pct_formula(2), 45, 'v1.0', 'тестовая заглушка — skyeng.ru'])
ws_runs.append(['RUN_002', '2026-05-02 14:30', 'foxford.ru', '62%',
                pct_formula(3), 52, 'v1.0', 'тестовая заглушка — foxford.ru'])

freeze_and_autofit(ws_runs)

# ====================================================
# ЛИСТ 2: Результаты
# ====================================================

ws_res = wb.create_sheet('Результаты')

style_header(ws_res, [
    'run_id', 'сайт', 'пункт_id', 'пункт_название',
    'результат_сканера', 'результат_эталон', 'сравнение',
    'тип_расхождения', 'серьёзность', 'комментарий'
])

# F: результат_эталон через VLOOKUP по ключу ключ_поиска листа Эталон
def f_etalon(r: int) -> str:
    return f'=IFERROR(VLOOKUP(B{r}&"|"&C{r},Эталон!$A:$E,5,0),"нет данных")'

# G: сравнение
def f_srav(r: int) -> str:
    e, f = f'E{r}', f'F{r}'
    return (
        f'=IF(OR({e}="не проверялось",{f}="нет данных",{f}=""),"НУЖНА ПРОВЕРКА",'
        f'IF(OR('
        f'AND({e}="PASS",OR({f}="Да",{f}="Частично(неявное)",{f}="Не обнаружен")),'
        f'AND({e}="FAIL",OR({f}="Нет",{f}="Обнаружен"))),'
        f'"✓","РАСХОЖДЕНИЕ"))'
    )


# 5 тестовых строк RUN_001 skyeng.ru + 5 строк RUN_002 foxford.ru
SAMPLE_ROWS = [
    # run_id, сайт, id, название, результат_сканера, disc_type, severity, comment
    ('RUN_001', 'skyeng.ru', '1',  'Согласие на обработку ПД в форме',    'PASS',            '', '', ''),
    ('RUN_001', 'skyeng.ru', '2',  'Чекбокс согласия не предотмечен',     'PASS',            '', '', ''),
    ('RUN_001', 'skyeng.ru', '3',  'Ссылка на политику рядом с формой',   'PASS',            '', '', ''),
    ('RUN_001', 'skyeng.ru', '4',  'Отдельный чекбокс для маркетинга',    'FAIL',            'false negative', 'критично', 'Эталон: Да, сканер пропустил'),
    ('RUN_001', 'skyeng.ru', '5',  'Минимальность собираемых данных',     'PASS',            '', '', ''),
    ('RUN_002', 'foxford.ru', '1', 'Согласие на обработку ПД в форме',    'PASS',            '', '', ''),
    ('RUN_002', 'foxford.ru', '2', 'Чекбокс согласия не предотмечен',     'PASS',            'false positive', 'важно', 'Эталон: Нет, сканер ложно нашёл'),
    ('RUN_002', 'foxford.ru', '3', 'Ссылка на политику рядом с формой',   'PASS',            '', '', ''),
    ('RUN_002', 'foxford.ru', '4', 'Отдельный чекбокс для маркетинга',    'FAIL',            '', '', ''),
    ('RUN_002', 'foxford.ru', '5', 'Минимальность собираемых данных',     'не проверялось',  '', '', ''),
]

for run_id, site, iid, iname, result, disc, sev, comm in SAMPLE_ROWS:
    ws_res.append([run_id, site, iid, iname, result, '', '', disc, sev, comm])

# Проставляем формулы F и G для всех строк с данными
for rn in range(2, ws_res.max_row + 1):
    ws_res[f'F{rn}'].value = f_etalon(rn)
    ws_res[f'G{rn}'].value = f_srav(rn)

# Условное форматирование — колонка G (сравнение)
cond_fmt(ws_res, 'G2:G10000', 'G2="РАСХОЖДЕНИЕ"', 'FFC7CE')
cond_fmt(ws_res, 'G2:G10000', 'G2="✓"',           'C6EFCE')
cond_fmt(ws_res, 'G2:G10000', 'G2="НУЖНА ПРОВЕРКА"', 'FFEB9C')

# Условное форматирование — колонка I (серьёзность)
cond_fmt(ws_res, 'I2:I10000', 'I2="критично"', 'FFC7CE')
cond_fmt(ws_res, 'I2:I10000', 'I2="важно"',    'FFDCB4')
cond_fmt(ws_res, 'I2:I10000', 'I2="минор"',    'FFEB9C')

freeze_and_autofit(ws_res)

# ====================================================
# ЛИСТ 3: Задачи
# ====================================================

ws_tasks = wb.create_sheet('Задачи')

style_header(ws_tasks, [
    'task_id', 'run_id', 'сайт', 'пункт_id', 'пункт_название',
    'тип_расхождения', 'серьёзность', 'описание_бага',
    'гипотеза_причины', 'статус', 'дата_создания',
    'дата_закрытия', 'комментарий_после_фикса'
])

SAMPLE_TASKS = [
    [
        'BUG_001', 'RUN_001', 'skyeng.ru', '4', 'Отдельный чекбокс для маркетинга',
        'false negative', 'критично',
        'Сканер выдал FAIL, эталон ожидает Да (маркетинговый чекбокс присутствует). Сайт: skyeng.ru',
        'Сканер не распознаёт признак нарушения. Возможно: неверный селектор, '
        'логика проверки не покрывает этот кейс',
        'открыта', '2026-05-01', '', ''
    ],
    [
        'BUG_002', 'RUN_002', 'foxford.ru', '2', 'Чекбокс согласия не предотмечен',
        'false positive', 'важно',
        'Сканер выдал PASS, эталон ожидает Нет (чекбокс предотмечен — нарушение). Сайт: foxford.ru',
        'Сканер ложно срабатывает. Возможно: слишком широкое условие, '
        'неверная интерпретация элемента',
        'открыта', '2026-05-02', '', ''
    ],
]

for t in SAMPLE_TASKS:
    ws_tasks.append(t)

cond_fmt(ws_tasks, 'G2:G10000', 'G2="критично"', 'FFC7CE')
cond_fmt(ws_tasks, 'G2:G10000', 'G2="важно"',    'FFDCB4')
cond_fmt(ws_tasks, 'G2:G10000', 'G2="минор"',    'FFEB9C')

freeze_and_autofit(ws_tasks)

# ====================================================
# ЛИСТ 4: Эталон
# ====================================================

ws_etalon = wb.create_sheet('Эталон')

style_header(ws_etalon, [
    'ключ_поиска', 'сайт', 'пункт_id', 'пункт_название',
    'эталонный_результат', 'дата_проверки', 'проверял'
])
# Пометим колонку-ключ серым (она служебная, нужна для VLOOKUP)
ws_etalon['A1'].fill = LGRAY_FILL

CHECK_DATE = '2025-04-27'
rows_added = 0

for col_i, site, ann in site_cols:
    for item_id, item_name, _group in checklist:
        val = ref_data.get((site, item_id))
        if val:
            ws_etalon.append([
                f'{site}|{item_id}',  # ключ для VLOOKUP
                site,
                item_id,
                item_name,
                val,
                CHECK_DATE,
                ann
            ])
            rows_added += 1

freeze_and_autofit(ws_etalon, narrow_a=True)

# =============================================
# ШАГ 4: Сохраняем
# =============================================

out_path = 'scanner_test_log.xlsx'
wb.save(out_path)

print(f'✓ Файл {out_path} создан')
print(f'  Пунктов чеклиста:  {len(checklist)}')
print(f'  Уникальных сайтов: {len(site_cols)}')
print(f'  Строк в Эталоне:   {rows_added}')
print(f'  Прогонов (заглушки): 2  (RUN_001 skyeng.ru, RUN_002 foxford.ru)')
print()
print('=== КАК ДОБАВИТЬ НОВЫЙ ПРОГОН ===')
print()
print('1. На лист "Прогоны" добавь строку:')
print('   run_id=RUN_00X, дата_время=ГГГГ-ММ-ДД ЧЧ:ММ, сайт=https://...,')
print('   score_сканера=XX%, время_сек=XX, версия_сканера=..., примечания=...')
print('   (колонка совпадений_с_эталоном_% заполнится формулой автоматически)')
print()
print('2. На лист "Результаты" добавь по одной строке на каждый пункт чеклиста:')
print('   - run_id         — тот же, что в шаге 1')
print('   - сайт           — тот же сайт')
print('   - пункт_id       — номер из листа "Эталон" (1–37)')
print('   - пункт_название — название из листа "Эталон"')
print('   - результат_сканера — PASS / FAIL / не проверялось')
print('   - результат_эталон и сравнение — заполняются формулами автоматически')
print('   - тип_расхождения / серьёзность — только если "сравнение"=РАСХОЖДЕНИЕ')
print()
print('3. Для пунктов, которых нет в выводе сканера — пиши результат_сканера="не проверялось"')
print()
print('4. Запусти: python scripts/recalc.py scanner_test_log.xlsx')
